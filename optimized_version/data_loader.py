"""
数据加载模块：从Excel读取所有基础数据
"""
import pandas as pd
import numpy as np
from openpyxl import load_workbook


def load_all_data(filepath='data.xlsx'):
    wb = load_workbook(filepath, read_only=True)

    # ── 1. 设备通用性 ──────────────────────────────────────────────
    ws = wb['设备通用性']
    rows = [r for r in ws.iter_rows(values_only=True) if any(v is not None for v in r)]
    # 第2行是标题：工序, 设备, 01..15
    header = rows[1]
    product_types = [str(h) for h in header[2:]]   # ['01'..'15']
    device_compatibility = {}   # {device_id: set of product_types}
    device_to_stage = {}        # {device_id: stage}
    stage_devices = {}          # {stage: [device_ids]}

    for row in rows[2:]:
        stage = str(row[0])
        device = str(row[1])
        device_to_stage[device] = stage
        stage_devices.setdefault(stage, []).append(device)
        compatible = set()
        for i, val in enumerate(row[2:]):
            if val == 1:
                compatible.add(product_types[i])
        device_compatibility[device] = compatible

    # ── 2. 生产加工时间 ─────────────────────────────────────────────
    ws = wb['生产加工时间']
    rows = [r for r in ws.iter_rows(values_only=True) if any(v is not None for v in r)]
    processing_time = {}   # {(device, product_type): seconds}

    for row in rows[2:]:
        stage = str(row[0])
        device = str(row[1])
        for i, val in enumerate(row[2:]):
            if val is not None:
                processing_time[(device, product_types[i])] = int(val)

    # ── 3. 换料时间 ──────────────────────────────────────────────
    # 规则：B工序所有型号切换=45min，D工序=20min
    # E工序：01-05 / 06-08 / 09-15 组内不换料，跨组换料=20min
    setup_time = {}    # {stage: setup_seconds or dict}
    ws = wb['换料时间']
    rows = [r for r in ws.iter_rows(values_only=True) if any(v is not None for v in r)]
    for row in rows[1:]:
        stage = str(row[0])
        minutes = int(row[2])
        setup_time[stage] = minutes * 60

    # E工序特殊处理：分组
    E_GROUPS = {
        'G1': set(['01', '02', '03', '04', '05']),
        'G2': set(['06', '07', '08']),
        'G3': set(['09', '10', '11', '12', '13', '14', '15']),
    }
    setup_time['E_groups'] = E_GROUPS

    # ── 4. 订单表 ────────────────────────────────────────────────
    ws = wb['订单']
    orders = []
    rows = [r for r in ws.iter_rows(values_only=True) if any(v is not None for v in r)]
    for row in rows[1:]:
        orders.append({
            'order_id': str(row[0]),
            'product_type': str(row[1]),
            'quantity': int(row[2]),
            'order_type': str(row[3]),           # MTO / MTS
            'due_time': float(row[4]) * 60,      # 转为秒
        })

    # ── 5. 最短剩余加工时间 ───────────────────────────────────────
    ws = wb['最短剩余加工时间']
    rows = [r for r in ws.iter_rows(values_only=True) if any(v is not None for v in r)]
    min_remaining = {}   # {order_id: {stage: seconds}}
    stage_cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
    for row in rows[1:]:
        oid = str(row[0])
        min_remaining[oid] = {s: int(row[i+1]) for i, s in enumerate(stage_cols)}

    return {
        'product_types': product_types,
        'device_compatibility': device_compatibility,
        'device_to_stage': device_to_stage,
        'stage_devices': stage_devices,
        'processing_time': processing_time,
        'setup_time': setup_time,
        'orders': orders,
        'min_remaining': min_remaining,
    }


def get_setup_seconds(data, stage, from_type, to_type, idle_time):
    """返回实际需要的换料时间（秒）。若空闲已够则返回0。"""
    if from_type is None or from_type == to_type:
        return 0
    if stage == 'B':
        needed = data['setup_time'].get('B', 0)
    elif stage == 'D':
        needed = data['setup_time'].get('D', 0)
    elif stage == 'E':
        groups = data['setup_time']['E_groups']
        from_grp = next((g for g, s in groups.items() if from_type in s), None)
        to_grp = next((g for g, s in groups.items() if to_type in s), None)
        needed = data['setup_time'].get('E', 0) if from_grp != to_grp else 0
    else:
        needed = 0
    return 0 if idle_time >= needed else needed


if __name__ == '__main__':
    data = load_all_data('/mnt/user-data/uploads/data.xlsx')
    print("订单数:", len(data['orders']))
    print("设备数:", len(data['device_compatibility']))
    print("加工时间条目:", len(data['processing_time']))
    print("示例订单:", data['orders'][0])
