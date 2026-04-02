"""
数据加载模块：从 Excel 中读取所有配套数据
"""
import pandas as pd
import numpy as np
from openpyxl import load_workbook

EXCEL_PATH = "/kaggle/working/PPO/data.xlsx"

# 产品型号列表（字符串）
PRODUCT_TYPES = [f"{i:02d}" for i in range(1, 16)]   # '01'~'15'
N_TYPES = 15

# 工序顺序
OPS = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
OPS_IDX = {op: i for i, op in enumerate(OPS)}

# B/C 并行，D 需要等 A 和 (B+C) 都完成
# 前序依赖：op -> list of ops that must finish first
PREREQS = {
    'A': [],
    'B': [],
    'C': ['B'],   # B 必须先于 C（或 B、C 并行，但 C 不依赖 B——见下方说明）
    'D': ['A', 'C'],
    'E': ['D'],
    'F': ['E'],
    'G': ['F'],
}
# 注：PDF 中 B 和 C 是并行链路上的串行（B→C），A 与 (B→C) 并行，D 等待两者都完成

def load_all():
    wb = load_workbook(EXCEL_PATH, read_only=True)

    # ---- 设备通用性 ----
    ws = wb['设备通用性']
    rows = list(ws.iter_rows(values_only=True))
    # row0: 合并标题, row1: 列头
    header = rows[1]   # ('工序','设备','01',...,'15')
    compatibility = {}   # {machine_id: set of product_type_str}
    machine_op = {}      # {machine_id: op}
    ops_machines = {op: [] for op in OPS}
    for row in rows[2:]:
        if row[0] is None and row[1] is None:
            continue
        op, machine = row[0], row[1]
        if op is None or machine is None:
            continue
        can_process = set()
        for j, pt in enumerate(PRODUCT_TYPES):
            val = row[2 + j]
            if val == 1:
                can_process.add(pt)
        compatibility[machine] = can_process
        machine_op[machine] = op
        ops_machines[op].append(machine)

    # ---- 生产加工时间 ----
    ws = wb['生产加工时间']
    rows = list(ws.iter_rows(values_only=True))
    proc_time = {}   # {(machine_id, product_type): seconds}
    for row in rows[2:]:
        if row[0] is None and row[1] is None:
            continue
        op, machine = row[0], row[1]
        if op is None or machine is None:
            continue
        for j, pt in enumerate(PRODUCT_TYPES):
            val = row[2 + j]
            if val is not None:
                proc_time[(machine, pt)] = int(val)

    # ---- 换料时间 ----
    ws = wb['换料时间']
    rows = list(ws.iter_rows(values_only=True))
    # 结构: (工序, 描述, 换料时间/分钟)
    setup_time = {}  # {op: {'time_min': int, 'groups': list_of_sets or None}}
    for row in rows[1:]:
        if row[0] is None:
            continue
        op, desc, t = row[0], row[1], row[2]
        if op == 'E':
            # 组间切换：01-05, 06-08, 09-15
            groups = [
                {'01','02','03','04','05'},
                {'06','07','08'},
                {'09','10','11','12','13','14','15'},
            ]
            setup_time[op] = {'time_min': int(t), 'groups': groups}
        else:
            setup_time[op] = {'time_min': int(t), 'groups': None}

    # ---- 订单 ----
    ws = wb['订单']
    rows = list(ws.iter_rows(values_only=True))
    orders = []
    for row in rows[1:]:
        if row[0] is None:
            continue
        oid, ptype, qty, mode, due = row
        orders.append({
            'order_id': oid,
            'product_type': str(ptype).zfill(2),
            'quantity': int(qty),
            'mode': mode,           # 'MTO' or 'MTS'
            'due_time': int(due),   # minutes
        })

    # ---- 最短剩余加工时间 ----
    ws = wb['最短剩余加工时间']
    rows = list(ws.iter_rows(values_only=True))
    min_remaining = {}   # {order_id: {op: seconds}}
    op_col = ['A','B','C','D','E','F','G']
    for row in rows[1:]:
        if row[0] is None:
            continue
        oid = row[0]
        min_remaining[oid] = {op: row[1+i] for i, op in enumerate(op_col)}

    wb.close()
    return {
        'compatibility': compatibility,
        'machine_op': machine_op,
        'ops_machines': ops_machines,
        'proc_time': proc_time,
        'setup_time': setup_time,
        'orders': orders,
        'min_remaining': min_remaining,
    }


if __name__ == '__main__':
    data = load_all()
    print("ops_machines:", {k: v for k, v in data['ops_machines'].items()})
    print("orders count:", len(data['orders']))
    print("setup_time:", data['setup_time'])
    print("sample proc_time (A1,08):", data['proc_time'].get(('A1','08')))
